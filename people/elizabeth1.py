from elizabeth import *
from openpyxl import *
from random import *


p = Personal('ru')
a = Address('ru')
f = Food('ru')
d = Datetime('ru')
wb = Workbook()
ws = wb.active
ws.cell(row=1, column=1).value = '##'
ws.cell(row=1, column=2).value = 'Имя'
ws.cell(row=1, column=3).value = 'Фамилия'
ws.cell(row=1, column=4).value = 'Пол'
ws.cell(row=1, column=5).value = 'Проффесия'
ws.cell(row=1, column=6).value = 'Город'
ws.cell(row=1, column=7).value = 'Любимое блюдо'
ws.cell(row=1, column=8).value = 'Адрес'
ws.cell(row=1, column=9).value = 'Телефон'
ws.cell(row=1, column=10).value = 'Email'
ws.cell(row=1, column=11).value = 'Дата рождения'
names = {}
ns = []
for i in range(2, 1002):
    g = randint(0, 1)
    t = ''
    if g == 0:
        g = 'male'
        t = 'Мужчина'
    else:
        g = 'female'
        t = 'Женщина'
    n = p.name(gender=g)
    s = p.surname()
    if n not in names:
        names[n] = 0
    names[n] += 1
    ns.append((n, s))
    ws.cell(row=i, column=1).value = i - 1
    ws.cell(row=i, column=2).value = n
    ws.cell(row=i, column=3).value = s
    ws.cell(row=i, column=4).value = t
    ws.cell(row=i, column=5).value = p.occupation()
    ws.cell(row=i, column=6).value = a.city()
    ws.cell(row=i, column=7).value = f.dish()
    ws.cell(row=i, column=8).value = a.address()
    ws.cell(row=i, column=9).value = p.telephone()
    ws.cell(row=i, column=10).value = p.email(gender=g)
    ws.cell(row=i, column=11).value = d.date(1938, 2017)
mx = -1
maxname = ''
for i in names:
    if names[i] > mx:
        mx = names[i]
        maxname = i
# print(names)
for i in range(len(ns)):
    if ns[i][0] == maxname:
        print(*ns[i])
wb.save("elizabeth1.xlsx")